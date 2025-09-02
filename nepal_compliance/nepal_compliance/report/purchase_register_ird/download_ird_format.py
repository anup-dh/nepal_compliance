# openpyxl 3.1.5 - MIT License, (c) 2024, Eric Gazoni, Charlie Clark, See License at https://pypi.org/project/openpyxl/

import frappe
from frappe import _
import json
import openpyxl
from frappe.utils import get_site_path
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

def convert_to_nepali_fy_format(year_start_date, year_end_date):
    try:
        start_year = year_start_date.year
        end_year = year_end_date.year

        nep_start = start_year + 57
        nep_end = end_year + 57

        return f"{nep_start}/{str(nep_end)[-2:]}"
    except Exception as e:
        return f"{year_start_date.year}-{year_end_date.year}"

@frappe.whitelist()
def generate_ird_purchase_register_excel():
    from nepal_compliance.nepal_compliance.report.purchase_register_ird.purchase_register_ird import get_data

    filters = frappe._dict(json.loads(frappe.form_dict.get("filters") or "{}"))
    rows = get_data(filters)

    if not rows:
        frappe.throw(_("No data found for the selected filters."))

    company = filters.get("company")
    company_info = frappe.get_doc("Company", company) if company else None
    company_name = company_info.company_name if company_info else "Company Name"
    # address = frappe.db.get_value("Address", {"is_your_company_address": 1}, "address_line1") or ""
    pan = company_info.tax_id or "N/A"
    invoice_name = frappe.db.get_value("Purchase Invoice", {"bill_no": rows[0].get("invoice")}, "name") or rows[0].get("invoice")
    try:
        invoice_doc = frappe.get_doc("Purchase Invoice", invoice_name)
        posting_date = invoice_doc.posting_date
    except frappe.DoesNotExistError:
        frappe.throw(_("Purchase Invoice {0} not found").format(invoice_name))

    fy_rows = frappe.db.get_all(
        "Fiscal Year",
        filters={
            "year_start_date": ["<=", posting_date],
            "year_end_date": [">=", posting_date],
        },
        fields=["name", "year_start_date", "year_end_date"],
        order_by="year_start_date desc",
        limit=1,
    )
    if not fy_rows:
        frappe.throw(_("No Fiscal Year found covering posting date {0}.").format(posting_date))
    fy = fy_rows[0]
    year_start_date = fy["year_start_date"]
    year_end_date = fy["year_end_date"]

    fiscal_year_nepali = convert_to_nepali_fy_format(year_start_date, year_end_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Register"

    bold_center = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def format_cell(cell):
        cell.alignment = center
        cell.font = bold_center
        cell.border = border

    ws.merge_cells("A1:M1")
    ws["A1"] = f"खरिद खाता"
    ws["A1"].font = Font(bold=True, size=16)

    ws.merge_cells("A2:M2")
    ws["A2"] = f"(नियम २३ को उपनियम (१) को खण्ड  (छ) संग सम्बन्धित )"
    ws["A2"].font = Font(bold=False)

    ws.merge_cells("A3:M3")
    ws["A3"] = ""

    ws.merge_cells("A4:M4")
    ws["A4"] = f"करदाता दर्ता नं (PAN): {pan}        करदाताको नाम: {company_name}         आर्थिक वर्ष: {fiscal_year_nepali}"
    ws["A4"].font = bold_center
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r in range(1, 5):
        ws[f"A{r}"].font = bold_center
        ws[f"A{r}"].alignment = center

    top_headers = [
        ("बीजक", 5),
        ("जम्मा खरिद मूल्य (रु)", 1),
        ("कर छुट हुने वस्तु वा सेवाको खरिद / पैठारी मूल्य (रु)", 1),
        ("करयोग्य खरिद (पूंजीगत बाहेक)", 2),
        ("करयोग्य पैठारी (पूंजीगत बाहेक)", 2),
        ("पूंजीगत करयोग्य खरिद / पैठारी", 2)
    ]

    col = 1
    for title, span in top_headers:
        if span == 1:
            ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        else:
            ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + span - 1)
        cell = ws.cell(row=5, column=col, value=title)
        format_cell(cell)
        col += span

    sub_headers = {
        1: "मिति",
        2: "बीजक नं.",
        3: "प्रज्ञापनपत्र नं.",
        4: "आपूर्तिकर्ताको नाम",
        5: "आपूर्तिकर्ताको स्थायी लेखा नम्बर",
        6: "मूल्य (रु)",
        7: "मूल्य (रु)",
        8: "मूल्य (रु)",
        9: "कर (रु)",
        10: "मूल्य (रु)",
        11: "कर (रु)",
        12: "मूल्य (रु)",
        13: "कर (रु)"
    }

    for col_num in range(1, 14):
        value = sub_headers.get(col_num)
        if not value:
            continue

        coord = ws.cell(row=6, column=col_num).coordinate
        is_top_left = True
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range and coord != merged_range.start_cell.coordinate:
                is_top_left = False
                break

        if is_top_left:
            cell = ws.cell(row=6, column=col_num, value=value)
            format_cell(cell)

    for row_idx, inv in enumerate(rows, start=7):
        row_data = [
            inv.get("nepali_date"),
            inv.get("invoice"),
            inv.get("customs_declaration_number"),
            inv.get("supplier_name"),
            inv.get("pan"),
            inv.get("total"),
            inv.get("tax_exempt"),
            inv.get("taxable_amount"),
            inv.get("tax_amount"),
            inv.get("taxable_import_non_capital_amount"),
            inv.get("taxable_import_non_capital_tax"),
            inv.get("capital_taxable_amount"),
            inv.get("capital_taxable_tax")
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    total_row = len(rows) + 7
    ws.cell(row=total_row, column=1, value="Total")
    format_cell(ws.cell(row=total_row, column=1))
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)

    for col in range(1, 14):
        if col == 5:
            continue
        col_letter = get_column_letter(col)
        col_total = f"=SUM({col_letter}7:{col_letter}{total_row - 1})"
        if not any([coord in merged_range for merged_range in ws.merged_cells.ranges for coord in [ws.cell(row=total_row, column=col).coordinate]]):
            cell = ws.cell(row=total_row, column=col, value=col_total)
            cell.border = border
            cell.alignment = center
            cell.number_format = '#,##0.00'

    for row_idx in range(7, total_row):
        ws.cell(row=row_idx, column=5).font = Font(color="000000")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

    file_name = "IRD_Purchase_Register.xlsx"
    path = get_site_path("public", "files", file_name)
    wb.save(path)

    return f"/files/{file_name}"
